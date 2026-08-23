"""Genera las fixtures .xlsx. Ver README.md: necesita openpyxl y xlsxwriter, que no son
dependencias del proyecto y se instalan en un entorno aparte."""

import sys
from openpyxl import Workbook
import xlsxwriter

out = sys.argv[1]

# 1. Libro de una sola hoja, escrito por openpyxl (cadenas en línea, `t="inlineStr"`).
wb = Workbook()
ws = wb.active
ws.title = "Vocabulario"
for row in [["Front", "Back"], ["Hello", "Hola"], ["House", "Casa"], ["Tree", "Árbol"]]:
    ws.append(row)
wb.save(f"{out}/basico.xlsx")

# 2. Libro de varias hojas: una sin tabla, dos con datos y encabezados distintos.
wb = Workbook()
ws = wb.active
ws.title = "Instrucciones"
ws.append(["Rellena las hojas siguientes y guarda el archivo."])

ingles = wb.create_sheet("Inglés")
for row in [["Question", "Answer"], ["Capital de Francia", "París"], ["2+2", "4"]]:
    ingles.append(row)

historia = wb.create_sheet("Historia")
for row in [["Columna A", "Columna B"], ["1492", "Llegada a América"], ["1789", "Revolución francesa"]]:
    historia.append(row)

wb.save(f"{out}/multihoja.xlsx")

# 3. Libro escrito por XlsxWriter: usa tabla de cadenas compartidas (`t="s"`), no en línea.
#    Cubre la otra forma en que un .xlsx real puede guardar su texto.
book = xlsxwriter.Workbook(f"{out}/compartidas.xlsx")
sheet = book.add_worksheet("Términos")
datos = [
    ["Término", "Definición"],
    ["Ñandú", "Ave corredora sudamericana"],
    ["Tom & Jerry", "Dibujos <animados>"],
    ["Árbol", "Planta perenne de tronco leñoso"],
]
for r, row in enumerate(datos):
    for c, value in enumerate(row):
        sheet.write(r, c, value)
book.close()
print("fixtures xlsx generadas")
